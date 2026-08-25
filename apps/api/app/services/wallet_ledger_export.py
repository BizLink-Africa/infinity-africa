"""Builds the Wallet Ledger .xlsx workbook for
GET /v1/merchant/wallet/ledger/export — see app/routers/merchant_portal.py.

Takes the same row shape app/services/ledger.py::export_wallet_ledger_rows
returns (one dict per ledger_entries row, newest first, already date-
filtered and joined onto its transaction) — no query logic here, purely
presentation.
"""

import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.time import to_dar_es_salaam

_HEADER = [
    "Merchant ID",
    "Business Name",
    "Date",
    "Transaction ID",
    "Type",
    "Direction",
    "Reference",
    "Provider Reference",
    "Payment Method",
    "Opening Balance",
    "Amount",
    "Charge / Fee",
    "Net Amount",
    "Closing Balance",
    "Status",
]

# Infinity Africa brand green (see apps/web/src/app/globals.css --color-primary).
_HEADER_FILL = PatternFill(start_color="04332A", end_color="04332A", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFFFF", bold=True)
_MONEY_FORMAT = "#,##0.00"


def _money(value: object) -> float | None:
    """Numeric cell, not a string — so Excel can sum/sort it. None stays
    None (not 0), matching the "Not available" convention used everywhere
    else in the ledger UI: an absent value is never fabricated as zero."""
    if value is None:
        return None
    return float(Decimal(str(value)))


def build_wallet_ledger_workbook(*, merchant: dict, rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Wallet Ledger"

    ws.append(_HEADER)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    merchant_code = merchant.get("merchant_code") or ""
    business_name = merchant.get("business_name") or ""

    money_columns = (10, 11, 12, 13, 14)  # Opening Balance .. Closing Balance, 1-indexed

    for row in rows:
        ws.append(
            [
                merchant_code,
                business_name,
                to_dar_es_salaam(row["date"]).replace(tzinfo=None),
                row.get("transaction_id") or "",
                row.get("type") or "",
                row["direction"],
                row.get("reference") or "",
                row.get("provider_reference") or "",
                row.get("method") or "",
                _money(row.get("balance_before")),
                _money(row.get("amount")),
                _money(row.get("fee_amount")),
                _money(row.get("net_amount")),
                _money(row.get("balance_after")),
                row.get("status") or "",
            ]
        )

    last_row = ws.max_row
    for col_idx in money_columns:
        letter = get_column_letter(col_idx)
        for r in range(2, last_row + 1):
            ws[f"{letter}{r}"].number_format = _MONEY_FORMAT

    date_letter = get_column_letter(3)
    for r in range(2, last_row + 1):
        ws[f"{date_letter}{r}"].number_format = "yyyy-mm-dd hh:mm"

    # Auto-size: widest of header/content per column, capped so one long
    # reference/description doesn't blow out the whole sheet.
    for col_idx, header in enumerate(_HEADER, start=1):
        letter = get_column_letter(col_idx)
        widest = len(header)
        for cells in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            value = cells[0].value
            widest = max(widest, len(str(value)) if value is not None else 0)
        ws.column_dimensions[letter].width = min(widest + 2, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
